# -*- coding: utf-8 -*-
"""
MPS REPORTS - Exportacion a Excel (Fase 9)
Genera un .xlsx con openpyxl con multiples pestanas.
Cada pestaña tiene encabezado con marca, modulo, fecha/hora y usuario.
"""

from datetime import datetime
import os

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                 GradientFill)
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# Colores corporativos en HEX sin #
AZUL_CORP  = "1E6FB8"
AZUL2_CORP = "2E89D6"
BLANCO     = "FFFFFF"
GRIS_CLARO = "EEF2F6"
GRIS_BORDE = "B9CCDD"
ROJO_CORP  = "D14545"
VERDE_CORP = "1E9E6A"
TEXTO_OSC  = "1A2A38"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=TEXTO_OSC, size=10, italic=False):
    return Font(name="Segoe UI", bold=bold, color=color, size=size,
                italic=italic)


def _border():
    thin = Side(style="thin", color=GRIS_BORDE)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _escribir_encabezado(ws, empresa, modulo, usuario, fecha_hora, n_cols):
    """Escribe las 4 filas de encabezado comunes a todas las pestanas."""
    # Fila 1: nombre del sistema
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, "MPS REPORTS - " + empresa)
    c.font = _font(bold=True, color=BLANCO, size=14)
    c.fill = _fill(AZUL_CORP)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Fila 2: modulo
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=n_cols)
    c = ws.cell(2, 1, modulo.upper())
    c.font = _font(bold=True, color=BLANCO, size=11)
    c.fill = _fill(AZUL2_CORP)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Fila 3: fecha/hora y usuario
    ws.merge_cells(start_row=3, start_column=1,
                   end_row=3, end_column=n_cols)
    c = ws.cell(3, 1,
                "Generado: {}   |   Usuario: {}".format(fecha_hora, usuario))
    c.font = _font(italic=True, color="5A7286", size=9)
    c.fill = _fill(GRIS_CLARO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # Fila 4: separador
    ws.row_dimensions[4].height = 6


def _escribir_columnas(ws, fila_ini, columnas):
    """Escribe la fila de encabezados de columna con fondo azul."""
    for col_idx, titulo in enumerate(columnas, 1):
        c = ws.cell(fila_ini, col_idx, titulo)
        c.font = _font(bold=True, color=BLANCO, size=10)
        c.fill = _fill(AZUL_CORP)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = _border()
    ws.row_dimensions[fila_ini].height = 20


def _escribir_filas(ws, fila_ini, filas, n_cols):
    """Escribe las filas de datos con filas alternas."""
    for i, fila in enumerate(filas):
        color = BLANCO if i % 2 == 0 else "F1F6FA"
        for col_idx, valor in enumerate(fila, 1):
            c = ws.cell(fila_ini + i, col_idx, valor)
            c.font = _font(size=9)
            c.fill = _fill(color)
            c.border = _border()
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[fila_ini + i].height = 18


def _autoajustar(ws, n_cols, col_ini=1, min_w=10, max_w=50):
    for col in range(col_ini, col_ini + n_cols):
        ancho = min_w
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    ancho = max(ancho, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = ancho


def exportar(ruta, empresa, usuario, pestanas):
    """
    pestanas: lista de dicts con:
      { "nombre": str, "modulo": str,
        "columnas": [str,...], "filas": [[val,...], ...] }
    """
    if not OPENPYXL_OK:
        raise ImportError("openpyxl no esta instalado. "
                          "Instala con: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # quitar hoja vacia por defecto
    fecha_hora = datetime.now().strftime("%d/%m/%Y %H:%M")

    for p in pestanas:
        nombre_hoja = p["nombre"][:31]  # Excel limita a 31 chars
        ws = wb.create_sheet(title=nombre_hoja)
        cols = p["columnas"]
        filas = p["filas"]
        n_cols = max(len(cols), 1)

        _escribir_encabezado(ws, empresa, p["modulo"], usuario,
                             fecha_hora, n_cols)
        _escribir_columnas(ws, 5, cols)
        _escribir_filas(ws, 6, filas, n_cols)
        _autoajustar(ws, n_cols)

        # Freeze panes en fila 6 (debajo de encabezados)
        ws.freeze_panes = "A6"

        # Filtros automaticos sobre los datos
        if filas:
            ws.auto_filter.ref = "A5:{}{}".format(
                get_column_letter(n_cols), 5 + len(filas))

    wb.save(ruta)
    return ruta
